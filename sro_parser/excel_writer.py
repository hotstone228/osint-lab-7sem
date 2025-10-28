from openpyxl import Workbook


def create_sheet_name(base):
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in base)
    if not cleaned:
        cleaned = "sheet"
    return cleaned[:31]


def write_to_excel(datasets, path):
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for dataset in datasets:
        sheet_name = create_sheet_name(dataset["sheet"])
        worksheet = workbook.create_sheet(title=sheet_name)
        headers = dataset.get("headers", [])
        rows = dataset.get("rows", [])

        if headers:
            worksheet.append(headers)
        for row in rows:
            worksheet.append(row)

    workbook.save(path)
